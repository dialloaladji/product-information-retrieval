"""Batch evaluation script — runs all 50 GTINs through the pipeline and scores results."""
from __future__ import annotations

import csv
import json
import time
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path("/Users/aladjidiallo/Downloads/verified_gtin_dataset_50.xlsx")
API_URL = "http://localhost:8000/retrieve"
OUTPUT_CSV = Path("evaluation_results.csv")


def call_pipeline(gtin: str) -> tuple[dict, float]:
    body = json.dumps({"gtin": gtin}).encode()
    req = urllib.request.Request(
        API_URL,
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    t0 = time.perf_counter()
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read().decode())
    duration_ms = round((time.perf_counter() - t0) * 1000, 1)
    return result, duration_ms


def normalize(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower()


def field_match(expected: str | None, got: str | None) -> str:
    e = normalize(expected)
    g = normalize(got)
    if not e:
        return "N/A"
    if e == g:
        return "exact"
    if e in g or g in e:
        return "partial"
    return "miss"


def load_ground_truth() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def run_evaluation() -> None:
    ground_truth = load_ground_truth()
    print(f"Loaded {len(ground_truth)} GTINs from Excel\n")

    fieldnames = [
        "gtin",
        "confidence_score",
        "duration_ms",
        "manufacturer_match",
        "mpn_match",
        "product_name_match",
        "category_match",
        "expected_manufacturer",
        "got_manufacturer",
        "expected_mpn",
        "got_mpn",
        "expected_product_name",
        "got_product_name",
        "expected_category",
        "got_category",
        "error",
    ]

    results = []
    for i, gt in enumerate(ground_truth, 1):
        gtin = str(gt["gtin"]).strip()
        print(f"[{i:02d}/50] {gtin} ... ", end="", flush=True)
        try:
            pipeline_output, duration_ms = call_pipeline(gtin)
            row = {
                "gtin": gtin,
                "confidence_score": pipeline_output.get("confidence_score", 0.0),
                "duration_ms": duration_ms,
                "manufacturer_match": field_match(gt.get("manufacturer"), pipeline_output.get("manufacturer")),
                "mpn_match": field_match(gt.get("manufacturer_part_number"), pipeline_output.get("manufacturer_product_id")),
                "product_name_match": field_match(gt.get("product_name"), pipeline_output.get("product_name")),
                "category_match": field_match(gt.get("category"), pipeline_output.get("category")),
                "expected_manufacturer": gt.get("manufacturer") or "",
                "got_manufacturer": pipeline_output.get("manufacturer") or "",
                "expected_mpn": gt.get("manufacturer_part_number") or "",
                "got_mpn": pipeline_output.get("manufacturer_product_id") or "",
                "expected_product_name": gt.get("product_name") or "",
                "got_product_name": pipeline_output.get("product_name") or "",
                "expected_category": gt.get("category") or "",
                "got_category": pipeline_output.get("category") or "",
                "error": "",
            }
            print(f"confidence={row['confidence_score']:.2f}  mfr={row['manufacturer_match']}  mpn={row['mpn_match']}  {duration_ms}ms")
        except Exception as exc:
            print(f"ERROR: {exc}")
            row = {f: "" for f in fieldnames}
            row["gtin"] = gtin
            row["error"] = str(exc)
            row["duration_ms"] = 0

        results.append(row)

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    print(f"\nRésultats sauvegardés dans : {OUTPUT_CSV.resolve()}\n")
    print_summary(results)


def print_summary(results: list[dict]) -> None:
    total = len(results)
    errors = sum(1 for r in results if r["error"])
    ok = [r for r in results if not r["error"]]

    confidence_success = sum(1 for r in ok if float(r.get("confidence_score") or 0) > 0)

    def accuracy(field: str) -> str:
        valid = [r for r in ok if r[field] != "N/A"]
        if not valid:
            return "N/A"
        exact = sum(1 for r in valid if r[field] == "exact")
        partial = sum(1 for r in valid if r[field] == "partial")
        return f"{exact}/{len(valid)} exact  +{partial} partial  ({100*exact/len(valid):.0f}%)"

    avg_ms = round(sum(float(r["duration_ms"]) for r in ok) / len(ok)) if ok else 0

    print("=" * 55)
    print("RÉSUMÉ ÉVALUATION")
    print("=" * 55)
    print(f"Total GTINs          : {total}")
    print(f"Erreurs API          : {errors}")
    print(f"Confidence > 0       : {confidence_success}/{len(ok)}")
    print(f"Temps moyen          : {avg_ms} ms")
    print()
    print(f"Manufacturer         : {accuracy('manufacturer_match')}")
    print(f"MPN                  : {accuracy('mpn_match')}")
    print(f"Product name         : {accuracy('product_name_match')}")
    print(f"Category             : {accuracy('category_match')}")
    print("=" * 55)
    print(f"\nRapport complet : {OUTPUT_CSV.resolve()}")


if __name__ == "__main__":
    run_evaluation()
